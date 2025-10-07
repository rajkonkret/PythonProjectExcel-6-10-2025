import openpyxl

filename = 'video2.xlsx'

wb = openpyxl.load_workbook(filename)
ws = wb['vgsales']

ws['P1'] = 'Averages Sales'
ws['P2'] = '=AVERAGE(K2:K16328)'

wb.save('video2old.xlsx')
wb.close()

ws['Q1'] = "Number of populated cells"
ws['Q2'] = "=COUNTA(E2:E16328)"

wb.save('video2old.xlsx')
wb.close()

ws['S1'] = "Total Sports Sales"
ws['S2'] = '=COUNTIF(E2:E16328, "sports")'

wb.save('video2old.xlsx')
wb.close()

print(ws['S2'].value)  # wypisze formułe

ws['T1'] = "Total sum of sports sales"
ws['T2'] = '=SUMIF(E2:E16328, "Sports", K2:K16328)'

wb.save('video2old.xlsx')
wb.close()

ws['U1'] = "Rounded sum of Sports Sales"
ws['U2'] = '=CEILING(T2, 25)'

wb.save('video2old.xlsx')
wb.close()

ws['V1'] = "Rounded sum of Sports Sales"
ws['V2'] = '=CEILING(T2, 1)'

wb.save('video2old.xlsx')
wb.close()

ws['W1'] = "Rounded"
ws['W2'] = '=ROUND(T2, 0)'

wb.save('video2old.xlsx')
wb.close()

ws['X1'] = 'Floor'
ws['X2'] = '=FLOOR(T2, 25)'

wb.save('video2old.xlsx')
wb.close()

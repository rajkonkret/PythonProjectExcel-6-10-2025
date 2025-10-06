from pprint import pprint

import openpyxl

wb = openpyxl.load_workbook("../data/videogamesales.xlsx")
ws = wb.active

print(wb)  # <openpyxl.workbook.workbook.Workbook object at 0x0000017BB778D550>
print(ws)  # <Worksheet "vgsales">

ws = wb['vgsales']

print("Total number of rows:", ws.max_row)  # Total numbers of rows: 16328
print("Total number of columns:", ws.max_column)  # Total number of columns: 10

print("Value un cell A1 is:", ws['A1'].value)  # Value un cell A1 is: Rank

# wczytanie wszystkich kolumn z wiersza 1
# +1 - excel numeruje od 1, python od 0
# list comprehensions
value = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
print(value)

# ['Rank', 'Name', 'Platform', 'Year', 'Genre', 'Publisher', 'NA_Sales', 'EU_Sales', 'JP_Sales', 'Other_Sales']

# dane z kolumny 2
data = [ws.cell(row=i, column=2).value for i in range(2, 12)]  # od 2 do 11
print(data)
# ['Wii Sports',
#  'Super Mario Bros.',
#  'Mario Kart Wii',
#  'Wii Sports Resort',
#  'Pokemon Red/Pokemon Blue',
#  'Tetris',
#  'New Super Mario Bros.',
#  'Wii Play',
#  'New Super Mario Bros. Wii',
#  'Duck Hunt']

my_list = list()  # pusta lista
for value in ws.iter_rows(
        min_row=1,
        max_row=11,
        min_col=1,
        max_col=6,
        values_only=True
):
    my_list.append(value)

print(my_list)
pprint(my_list)
# [('Rank', 'Name', 'Platform', 'Year', 'Genre', 'Publisher'),
#  (1, 'Wii Sports', 'Wii', 2006, 'Sports', 'Nintendo'),
#  (2, 'Super Mario Bros.', 'NES', 1985, 'Platform', 'Nintendo'),
#  (3, 'Mario Kart Wii', 'Wii', 2008, 'Racing', 'Nintendo'),
#  (4, 'Wii Sports Resort', 'Wii', 2009, 'Sports', 'Nintendo'),
#  (5, 'Pokemon Red/Pokemon Blue', 'GB', 1996, 'Role-Playing', 'Nintendo'),
#  (6, 'Tetris', 'GB', 1989, 'Puzzle', 'Nintendo'),
#  (7, 'New Super Mario Bros.', 'DS', 2006, 'Platform', 'Nintendo'),
#  (8, 'Wii Play', 'Wii', 2006, 'Misc', 'Nintendo'),
#  (9, 'New Super Mario Bros. Wii', 'Wii', 2009, 'Platform', 'Nintendo'),
#  (10, 'Duck Hunt', 'NES', 1984, 'Shooter', 'Nintendo')]

for ele1, ele2, ele3, ele4, ele5, ele6 in my_list:
    # print(f"{ele1:<8}{ele2:>35}{ele3:^10}{ele4:<10}{ele5:<15}")
    print(f"{ele1:<8}{ele2:>35}{ele3:^10}{ele4:<10}{ele5:<15}{ele6:<15}")

# dopisanie nowej kolumny
ws['K1'] = "Sum of Sales"

wb.save('video2.xlsx')
wb.close()

import openpyxl

filename = 'video2.xlsx'

wb = openpyxl.load_workbook(filename)
ws = wb.active

row_position = 1

for i in range(1, ws.max_row):
    row_position += 1  # r = r + 1
    print(row_position)
    NA_Sales = ws.cell(row=row_position, column=7).value
    EU_Sales = ws.cell(row=row_position, column=8).value
    JP_Sales = ws.cell(row=row_position, column=9).value
    Other_Sales = ws.cell(row=row_position, column=10).value

    total_series = (NA_Sales + EU_Sales + JP_Sales + Other_Sales)
    ws.cell(row=row_position, column=11).value = total_series

# wykona się po zakończeniu pętli
wb.save(filename)
wb.close()
# https://www.kaggle.com/

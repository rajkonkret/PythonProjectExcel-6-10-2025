import openpyxl

from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# wb = openpyxl.load_workbook('video2.xlsx')
wb = openpyxl.load_workbook('video3.xlsx')
ws = wb.active

print(ws.title)  # vgsales

# # zmiana nazwy arkusza
# ws.title = "Video Games Sales Data"
#
# wb.save('video3.xlsx')
# wb.close()

# sheet_names = wb.sheetnames
# print(sheet_names)
# # ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year']
#
# # ustawienie aktywnego arkusza po indeksie
# sheet = wb[sheet_names[1]]  # indeks 1, czyli element 2
# wb.active = wb.index(sheet)

ws = wb.active
print(ws.title)  # Total Sales by Genre

# formatowanie tekstu w pierwszym wierszu
ws = wb['Video Games Sales Data']
ws['A1'].font = Font(bold=True, size=12)

for cell in ws[1:1]:
    cell.font = Font(bold=True, size=12)

wb.save('video3.xlsx')
wb.close()

# dodanie arkusza do pliku
# wb.create_sheet('Empty Sheet')
# print(wb.sheetnames)
# # ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year', 'Empty Sheet']
#
# wb.save('video3.xlsx')
# wb.close()

# usunięcie arkusza z pliku
# wb.remove(wb['Empty Sheet'])
# print(wb.sheetnames)
#
# wb.save('video3.xlsx')
# wb.close()
# ['Video Games Sales Data', 'Total Sales by Genre', 'Breakdown of Sales by Genre', 'Breakdown of Sales by Year']
# KeyError: 'Worksheet Empty Sheet does not exist.' - nie ma arkusza - błąd

# kopiowanie arkusza
wb.copy_worksheet(wb['Video Games Sales Data'])
wb.save('video3.xlsx')
wb.close()

import openpyxl

from openpyxl.styles import Font, colors, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# wb = openpyxl.load_workbook('video2.xlsx')
wb = openpyxl.load_workbook('video3.xlsx')
ws = wb.active

print(ws.title)  # vgsales

ws = wb['Video Games Sales Data']

# RGB -> FF0000 -> red
# https://htmlcolorcodes.com/
ws['A1'].font = Font(color='FF0000', bold=True, size=12)  # czerwony
# RGB -> 0000FF -> blue
ws['A2'].font = Font(color='0000FF')  # niebieski
# https://www.bing.com/search?q=color+codes+from+hex&cvid=f3aba924c36142478c9ea44805fec83a&gs_lcrp=EgRlZGdlKgYIABBFGDkyBggAEEUYOTIGCAEQABhAMgYIAhAAGEAyBggDEAAYQDIGCAQQABhAMgYIBRAAGEAyBwgGEOsHGEDSAQg3MzgyajBqNKgCCLACAQ&FORM=ANAB01&PC=U531

ws['A1'].fill = PatternFill('lightVertical', start_color='38e3ff')
# ws['A1'].fill = PatternFill('darkTrellis', start_color='1188A7')
# FILL_NONE = 'none'
# FILL_SOLID = 'solid'
# FILL_PATTERN_DARKDOWN = 'darkDown'
# FILL_PATTERN_DARKGRAY = 'darkGray'
# FILL_PATTERN_DARKGRID = 'darkGrid'
# FILL_PATTERN_DARKHORIZONTAL = 'darkHorizontal'
# FILL_PATTERN_DARKTRELLIS = 'darkTrellis'
# FILL_PATTERN_DARKUP = 'darkUp'
# FILL_PATTERN_DARKVERTICAL = 'darkVertical'
# FILL_PATTERN_GRAY0625 = 'gray0625'
# FILL_PATTERN_GRAY125 = 'gray125'
# FILL_PATTERN_LIGHTDOWN = 'lightDown'
# FILL_PATTERN_LIGHTGRAY = 'lightGray'
# FILL_PATTERN_LIGHTGRID = 'lightGrid'
# FILL_PATTERN_LIGHTHORIZONTAL = 'lightHorizontal'
# FILL_PATTERN_LIGHTTRELLIS = 'lightTrellis'
# FILL_PATTERN_LIGHTUP = 'lightUp'
# FILL_PATTERN_LIGHTVERTICAL = 'lightVertical'
# FILL_PATTERN_MEDIUMGRAY = 'mediumGray'

# ramka
# 'dashDot','dashDotDot', 'dashed','dotted',
#                             'double','hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
#                             'mediumDashed', 'slantDashDot', 'thick', 'thin'
# my_border = Side(border_style='thin', color='000000')  # czarny
my_border = Side(border_style='thick', color='000000')  # czarny
ws['A1'].border = Border(
    top=my_border, left=my_border, right=my_border, bottom=my_border
)

wb.save('video3.xlsx')
wb.close()
